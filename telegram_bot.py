import os
import logging
import openpyxl
from telegram import Update
from telegram.ext import ApplicationBuilder, ContextTypes, MessageHandler, filters
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# Configure logging
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# Ganti dengan token bot Telegram Anda
TOKEN = os.getenv('TELEGRAM_BOT_TOKEN', 'YOUR_BOT_TOKEN_HERE')

# Path ke file XLSX
XLSX_FILE = os.getenv('XLSX_FILE', 'data.xlsx')

def load_existing_data():
    """Memuat data dari XLSX jika file ada."""
    try:
        if os.path.exists(XLSX_FILE):
            wb = openpyxl.load_workbook(XLSX_FILE)
            sheet = wb.active
            return [row[0].value for row in sheet.iter_rows(min_row=2) if row[0].value]
        return []
    except Exception as e:
        logger.error(f"Error loading data: {e}")
        return []

def save_data(message):
    """Menyimpan pesan baru ke XLSX."""
    try:
        if not os.path.exists(XLSX_FILE):
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet['A1'] = 'Message'
        else:
            wb = openpyxl.load_workbook(XLSX_FILE)
            sheet = wb.active
        
        sheet.append([message])
        wb.save(XLSX_FILE)
        logger.info(f"Data saved: {message}")
    except Exception as e:
        logger.error(f"Error saving data: {e}")
        raise

async def message_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Handler untuk menangani pesan masuk."""
    try:
        user_message = update.message.text
        existing_data = load_existing_data()
        
        if user_message in existing_data:
            await update.message.reply_text("Pesan sudah ada di database.")
        else:
            await update.message.reply_text("Data dapat digunakan.")
            save_data(user_message)
    except Exception as e:
        logger.error(f"Error handling message: {e}")
        await update.message.reply_text("Terjadi kesalahan saat memproses pesan.")

def main():
    """Fungsi utama untuk menjalankan bot."""
    if TOKEN == 'YOUR_BOT_TOKEN_HERE':
        logger.error("Please set your Telegram bot token in .env file!")
        print("Error: Bot token not configured. Please create a .env file with TELEGRAM_BOT_TOKEN.")
        return
    
    try:
        application = ApplicationBuilder().token(TOKEN).build()
        application.add_handler(MessageHandler(filters.TEXT, message_handler))
        logger.info("Bot started successfully!")
        application.run_polling()
    except Exception as e:
        logger.error(f"Error starting bot: {e}")
        print(f"Error: {e}")

if __name__ == '__main__':
    main()